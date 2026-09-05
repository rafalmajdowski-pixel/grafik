from datetime import datetime, timedelta
import io
import math
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
import pulp
import streamlit as st

# --- KONFIGURACJA STRONY STREAMLIT ---
st.set_page_config(
    page_title="żabka jush! - Optymalizator Grafiku DS",
    page_icon="⚡",
    layout="wide",
)

# --- STYLIZACJA W PALECIE JUSH! ---
st.markdown(
    """
    <style>
    :root {
        --jush-lime: #8BC53F;
        --jush-dark-green: #005B2B;
        --jush-light-lime: #EBF7D4;
    }
    
    .stApp {
        background-color: #FAFCF5;
    }
    
    [data-testid="stSidebar"] {
        background-color: #8BC53F !important;
    }
    [data-testid="stSidebar"] * {
        color: #005B2B !important;
        font-weight: bold !important;
    }
    
    div.stButton > button {
        background-color: #005B2B !important;
        color: #8BC53F !important;
        font-weight: 800 !important;
        font-size: 16px !important;
        border-radius: 12px !important;
        border: none !important;
        padding: 10px 24px !important;
        transition: all 0.2s ease !important;
    }
    div.stButton > button:hover {
        background-color: #004420 !important;
        color: #A3DF52 !important;
    }
    
    /* Karty ustawień */
    .rule-card {
        background-color: white;
        border: 2px solid #8BC53F;
        border-radius: 12px;
        padding: 16px;
        margin-bottom: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    
    h1, h2, h3 {
        color: #005B2B !important;
        font-family: 'Arial Black', sans-serif !important;
    }
    </style>
""",
    unsafe_allow_html=True,
)

# --- BRANDING JUSH! - BANNER NAGŁÓWKA ---
col_logo, col_title = st.columns([1, 4])
with col_logo:
    st.image(
        "https://zabkagroup.com/wp-content/uploads/2022/09/Jush_logo.png",
        width=140,
    )
with col_title:
    st.markdown(
        "<h1 style='margin-bottom:0; font-size: 2.6rem;'>żabka <span style='color:#005B2B;'>jush!</span></h1>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p style='font-weight:bold; color:#005B2B; font-size: 1.1rem;'>Optymalizator Grafiku Pickerów DS</p>",
        unsafe_allow_html=True,
    )

st.divider()

# --- SIDEBAR: PARAMETRY EFEKTYWNOŚCI I OBSADY ---
st.sidebar.image(
    "https://zabkagroup.com/wp-content/uploads/2022/09/Jush_logo.png", width=110
)
st.sidebar.header("⚙️ Ustawienia Magazynu DS")

typ_magazynu = st.sidebar.selectbox("Typ magazynu", ["Standardowy", "Nocny"])
is_nocny = typ_magazynu == "Nocny"

godzina_otwarcia_ds = 6.0
godzina_zamkniecia_ds = 25.5 if is_nocny else 23.5
max_godzina_zamowien = 25 if is_nocny else 23

cel_efektywnosci = st.sidebar.number_input(
    "Efektywność pakowania (zamówienia / h / picker)",
    min_value=1,
    value=15,
)

min_zmiana = 6
max_zmiana = 12

MAPA_DNI = {
    "Monday": "Poniedziałek",
    "Tuesday": "Wtorek",
    "Wednesday": "Środa",
    "Thursday": "Czwartek",
    "Friday": "Piątek",
    "Saturday": "Sobota",
    "Sunday": "Niedziela",
}

# --- 1. ZAKRES DAT DLA GRAFIKU ---
st.header("1. Wybierz okres grafiku")
okres_grafiku = st.date_input(
    "Wskaż zakres od - do:",
    value=(datetime.now().date(), datetime.now().date() + timedelta(days=29)),
)

if isinstance(okres_grafiku, tuple) and len(okres_grafiku) == 2:
    start_date, end_date = okres_grafiku
    dni_zakresu = [
        start_date + timedelta(days=i)
        for i in range((end_date - start_date).days + 1)
    ]
else:
    dni_zakresu = [okres_grafiku[0]]

# --- 2. ANALIZA GODZINOWA Z LOOKERA ---
st.header("2. Wgraj raport zamówień z Lookera")
uploaded_file = st.file_uploader(
    "Wybierz plik CSV lub Excel z Lookera (hourly volume)", type=["csv", "xlsx"]
)

srednie_godzinowe = {}

if uploaded_file:
    try:
        if uploaded_file.name.endswith(".csv"):
            df_raw = pd.read_csv(uploaded_file)
        else:
            df_raw = pd.read_excel(uploaded_file)

        col_hour = None
        for c in df_raw.columns:
            if "hour" in str(c).lower() or "godz" in str(c).lower():
                col_hour = c
                break
        if not col_hour:
            col_hour = df_raw.columns[0]

        date_cols = {}
        for c in df_raw.columns:
            dt_val = pd.to_datetime(str(c).strip(), errors="coerce")
            if pd.notna(dt_val) and dt_val.year > 2020:
                dzien_nazwa = MAPA_DNI.get(
                    dt_val.strftime("%A"), dt_val.strftime("%A")
                )
                if dzien_nazwa not in date_cols:
                    date_cols[dzien_nazwa] = []
                date_cols[dzien_nazwa].append(c)

        godziny_data = {
            d: {h: [] for h in range(26)} for d in MAPA_DNI.values()
        }

        for idx, row in df_raw.iterrows():
            h_val = pd.to_numeric(row[col_hour], errors="coerce")
            if pd.notna(h_val) and 0 <= int(h_val) <= 25:
                h_int = int(h_val)
                for d_nazwa, cols_list in date_cols.items():
                    for c_date in cols_list:
                        val = pd.to_numeric(
                            str(row[c_date])
                            .replace(" ", "")
                            .replace(",", "."),
                            errors="coerce",
                        )
                        if pd.notna(val):
                            godziny_data[d_nazwa][h_int].append(val)

        for d_nazwa in MAPA_DNI.values():
            srednie_godzinowe[d_nazwa] = {}
            for h in range(26):
                vals = godziny_data[d_nazwa][h]
                sr_h = sum(vals) / len(vals) if vals else 0
                srednie_godzinowe[d_nazwa][h] = sr_h

        st.success("⚡ Raport Lookera przetworzony pomyślnie!")

    except Exception as e:
        st.error(f"Błąd odczytu pliku z Lookera: {e}")

# --- 3. PRACOWNICI & PRZEJRZYSTE USTAWIENIA ---
st.header("3. Zespół Pickerów DS & Indywidualne Reguły")

pracownicy_default = [
    "Aval01204VasinA",
    "Aval01209KushnY",
    "AvalZhukoD",
    "Dive01202VitalD",
    "Eter01203SavchV",
    "EterZaichI",
]
pracownicy_input = st.text_area(
    "Lista pickerów (każdy w nowej linii):",
    "\n".join(pracownicy_default),
)
pracownicy = [
    p.strip() for p in pracownicy_input.split("\n") if p.strip() != ""
]

# INICJALIZACJA SESSION STATE DLA ZASĄD INDYWIDUALNYCH
if "preferencje_dict" not in st.session_state:
    st.session_state.preferencje_dict = {}
if "korekty_godzin_dict" not in st.session_state:
    st.session_state.korekty_godzin_dict = {}
if "urlopy_list" not in st.session_state:
    st.session_state.urlopy_list = []

st.subheader("➕ Ustawienia dla Wybranego Pickera")

# OTWARTY, CZYSTY FORMULARZ (BEZ EXPANDERA)
col_sel, col_pref_type, col_val, col_add = st.columns([2, 3, 2, 2])

with col_sel:
    p_target = st.selectbox("Wybierz pickera:", pracownicy if pracownicy else ["-"])

with col_pref_type:
    type_opt = st.selectbox(
        "Rodzaj dodawanego ustawienia:",
        ["Preferencja Pory Dnia", "Modyfikacja Etatowa (+/- h)", "Nieobecność / Urlop (Całe Dni)"]
    )

with col_val:
    if type_opt == "Preferencja Pory Dnia":
        val_pref = st.selectbox("Pora dnia:", ["Preferuje Poranki (06:00)", "Preferuje Zamknięcia"])
    elif type_opt == "Modyfikacja Etatowa (+/- h)":
        val_hours = st.number_input("Różnica godzin (np. +20 lub -30):", value=0, step=5)
    else:
        val_dates = st.date_input("Zakres wolnego:", value=(datetime.now().date(), datetime.now().date()))

with col_add:
    st.write("&nbsp;")
    if st.button("➕ Dodaj regułę", use_container_width=True):
        if type_opt == "Preferencja Pory Dnia":
            st.session_state.preferencje_dict[p_target] = val_pref
            st.toast(f"Dodano preferencję dla {p_target}!", icon="🎯")
        elif type_opt == "Modyfikacja Etatowa (+/- h)":
            st.session_state.korekty_godzin_dict[p_target] = val_hours
            st.toast(f"Skorygowano etat dla {p_target} o {val_hours}h!", icon="⏱️")
        else:
            if isinstance(val_dates, tuple) and len(val_dates) == 2:
                st.session_state.urlopy_list.append({"Pracownik": p_target, "Od": val_dates[0], "Do": val_dates[1]})
            elif isinstance(val_dates, tuple) and len(val_dates) == 1:
                st.session_state.urlopy_list.append({"Pracownik": p_target, "Od": val_dates[0], "Do": val_dates[0]})
            st.toast(f"Zarejestrowano nieobecność dla {p_target}!", icon="📋")

# CZYTELNE TABELE ZAMIAST CZARNEGO JSON-A
st.write("---")
st.subheader("📋 Aktywne Ustawienia Zespołu:")

c_pref, c_kor, c_url = st.columns(3)

with c_pref:
    st.markdown("🎯 **Preferencje Pory Dnia**")
    if st.session_state.preferencje_dict:
        df_pref = pd.DataFrame(
            list(st.session_state.preferencje_dict.items()),
            columns=["Picker", "Preferowany Zwyczaj"]
        )
        st.dataframe(df_pref, use_container_width=True, hide_index=True)
        if st.button("🗑️ Wyczyść preferencje", key="c1"):
            st.session_state.preferencje_dict = {}
            st.rerun()
    else:
        st.caption("Brak ustalonych preferencji pory dnia.")

with c_kor:
    st.markdown("⏱️ **Korekty Etatów (+/- h)**")
    if st.session_state.korekty_godzin_dict:
        df_kor = pd.DataFrame(
            [{"Picker": k, "Zmiana Czasu": f"{v:+d}h"} for k, v in st.session_state.korekty_godzin_dict.items()]
        )
        st.dataframe(df_kor, use_container_width=True, hide_index=True)
        if st.button("🗑️ Wyczyść korekty", key="c2"):
            st.session_state.korekty_godzin_dict = {}
            st.rerun()
    else:
        st.caption("Wszyscy pickerzy mają domyślny etat.")

with c_url:
    st.markdown("🌴 **Nieobecności i Urlopy**")
    if st.session_state.urlopy_list:
        df_url = pd.DataFrame(st.session_state.urlopy_list)
        df_url["Od"] = df_url["Od"].apply(lambda x: x.strftime("%d/%m/%Y"))
        df_url["Do"] = df_url["Do"].apply(lambda x: x.strftime("%d/%m/%Y"))
        st.dataframe(df_url, use_container_width=True, hide_index=True)
        if st.button("🗑️ Wyczyść urlopy", key="c3"):
            st.session_state.urlopy_list = []
            st.rerun()
    else:
        st.caption("Brak nieobecności w grafiku.")

# --- 4. GENEROWANIE GRAFIKU Z LOGIKĄ JUSH! ---
st.divider()
st.header("4. Generowanie Grafiku Pickerów")
if st.button("🚀 Wygeneruj Grafik jush!", type="primary", use_container_width=True):
    if not uploaded_file:
        st.error("Proszę najpierw wgrać plik z Lookera!")
    elif not pracownicy:
        st.error("Proszę wpisać listę pickerów!")
    else:
        try:
            wymagani_pracownicy_h = {}
            total_required_hours = 0
            for d in dni_zakresu:
                d_nazwa = MAPA_DNI.get(d.strftime("%A"), d.strftime("%A"))
                wymagani_pracownicy_h[d] = {}
                for h in range(max_godzina_zamowien + 1):
                    sr_zam = srednie_godzinowe.get(d_nazwa, {}).get(h, 0)
                    potrzeba_osob = math.ceil(sr_zam / cel_efektywnosci)
                    wymagani_pracownicy_h[d][h] = potrzeba_osob
                    total_required_hours += potrzeba_osob

            model = pulp.LpProblem("Optymalizacja_Grafiku", pulp.LpMinimize)

            prawidlowe_zmiany = []
            starty = [6.0 + 0.5 * i for i in range(int((18.0 - 6.0) * 2) + 1)]
            dlugosci = [float(l) for l in range(min_zmiana, max_zmiana + 1)]

            for s in starty:
                for l in dlugosci:
                    koniec = s + l
                    if koniec <= godzina_zamkniecia_ds:
                        prawidlowe_zmiany.append((s, l))

            zmiany_ranne = [(s, l) for s, l in prawidlowe_zmiany if s == 6.0]
            zmiany_wieczorne = [
                (s, l)
                for s, l in prawidlowe_zmiany
                if (s + l) == godzina_zamkniecia_ds
            ]

            zmienne_zmian = []
            for p in pracownicy:
                for d in dni_zakresu:
                    for s, l in prawidlowe_zmiany:
                        zmienne_zmian.append((p, d, s, l))

            y = pulp.LpVariable.dicts("zmiana", zmienne_zmian, cat="Binary")

            work_day = pulp.LpVariable.dicts(
                "work_day",
                [(p, d) for p in pracownicy for d in dni_zakresu],
                cat="Binary",
            )

            penalty_rest_12h = pulp.LpVariable.dicts(
                "pen_rest",
                [(p, d) for p in pracownicy for d in dni_zakresu],
                lowBound=0,
                cat="Binary",
            )
            penalty_cadence = pulp.LpVariable.dicts(
                "pen_cadence",
                [(p, d) for p in pracownicy for d in dni_zakresu],
                lowBound=0,
                cat="Binary",
            )

            dev_plus = pulp.LpVariable.dicts(
                "dev_plus", pracownicy, lowBound=0, cat="Continuous"
            )
            dev_minus = pulp.LpVariable.dicts(
                "dev_minus", pracownicy, lowBound=0, cat="Continuous"
            )
            dev_ranne = pulp.LpVariable.dicts(
                "dev_ranne", pracownicy, lowBound=0, cat="Continuous"
            )
            dev_wieczor = pulp.LpVariable.dicts(
                "dev_wieczor", pracownicy, lowBound=0, cat="Continuous"
            )

            pref_penalty = []
            for p in pracownicy:
                pref = st.session_state.preferencje_dict.get(p, "Brak")
                for d in dni_zakresu:
                    if pref == "Preferuje Poranki (06:00)":
                        pref_penalty.append(
                            pulp.lpSum(y[p, d, s, l] for s, l in zmiany_wieczorne) * 10.0
                        )
                    elif pref == "Preferuje Zamknięcia":
                        pref_penalty.append(
                            pulp.lpSum(y[p, d, s, l] for s, l in zmiany_ranne) * 10.0
                        )

            model += (
                pulp.lpSum(
                    dev_plus[p]
                    + dev_minus[p]
                    + 2.0 * dev_ranne[p]
                    + 2.0 * dev_wieczor[p]
                    for p in pracownicy
                )
                + pulp.lpSum(
                    1000.0 * penalty_rest_12h[p, d]
                    + 500.0 * penalty_cadence[p, d]
                    for p in pracownicy
                    for d in dni_zakresu
                )
                + pulp.lpSum(pref_penalty)
            )

            for p in pracownicy:
                dni_absencji = sum(
                    1
                    for d in dni_zakresu
                    if any(
                        u["Pracownik"] == p and u["Od"] <= d <= u["Do"]
                        for u in st.session_state.urlopy_list
                    )
                )
                dni_dostepne = max(1, len(dni_zakresu) - dni_absencji)
                proporcja = dni_dostepne / len(dni_zakresu)
                
                korekta_h = st.session_state.korekty_godzin_dict.get(p, 0)
                target_p = ((total_required_hours / len(pracownicy)) * proporcja) + korekta_h

                suma_h_p = pulp.lpSum(
                    y[p, d, s, l] * l
                    for d in dni_zakresu
                    for s, l in prawidlowe_zmiany
                )

                model += suma_h_p <= target_p + 15.0
                model += suma_h_p >= target_p - 15.0
                model += suma_h_p + dev_minus[p] - dev_plus[p] == target_p

                num_ranne = pulp.lpSum(
                    y[p, d, s, l] for d in dni_zakresu for s, l in zmiany_ranne
                )
                num_wieczorne = pulp.lpSum(
                    y[p, d, s, l]
                    for d in dni_zakresu
                    for s, l in zmiany_wieczorne
                )

                model += num_ranne - num_wieczorne <= 3.0 + dev_ranne[p]
                model += num_wieczorne - num_ranne <= 3.0 + dev_wieczor[p]

                for idx_d, d in enumerate(dni_zakresu):
                    model += (
                        work_day[p, d]
                        == pulp.lpSum(y[p, d, s, l] for s, l in prawidlowe_zmiany)
                    )

                    for u in st.session_state.urlopy_list:
                        if u["Pracownik"] == p and u["Od"] <= d <= u["Do"]:
                            for s, l in prawidlowe_zmiany:
                                model += y[p, d, s, l] == 0

                    if idx_d < len(dni_zakresu) - 1:
                        d_next = dni_zakresu[idx_d + 1]
                        for s1, l1 in prawidlowe_zmiany:
                            koniec_d1 = s1 + l1
                            for s2, l2 in prawidlowe_zmiany:
                                start_d2 = s2 + 24.0
                                if (start_d2 - koniec_d1) < 12.0:
                                    model += (
                                        y[p, d, s1, l1] + y[p, d_next, s2, l2]
                                        <= 1 + penalty_rest_12h[p, d_next]
                                    )

                for idx_d in range(len(dni_zakresu) - 5):
                    window6 = [dni_zakresu[idx_d + i] for i in range(6)]
                    d_last = window6[-1]
                    model += (
                        pulp.lpSum(work_day[p, d_w] for d_w in window6)
                        <= 5 + 6 * penalty_cadence[p, d_last]
                    )

            for d in dni_zakresu:
                model += (
                    pulp.lpSum(
                        y[p, d, 6.0, l]
                        for p in pracownicy
                        for s, l in prawidlowe_zmiany
                        if s == 6.0
                    )
                    >= 1
                )
                model += (
                    pulp.lpSum(
                        y[p, d, s, l]
                        for p in pracownicy
                        for s, l in prawidlowe_zmiany
                        if s + l == godzina_zamkniecia_ds
                    )
                    >= 1
                )

                for h in range(7, max_godzina_zamowien):
                    potrzebni = wymagani_pracownicy_h[d].get(h, 0)
                    if potrzebni > 0:
                        pracujacy = [
                            y[p, d, s, l]
                            for p in pracownicy
                            for s, l in prawidlowe_zmiany
                            if s <= h and (s + l) >= (h + 1)
                        ]
                        model += pulp.lpSum(pracujacy) >= potrzebni

            status = model.solve(pulp.PULP_CBC_CMD(msg=False))

            st.session_state.schedule_generated = True
            st.session_state.pracownicy = pracownicy
            st.session_state.dni_zakresu = dni_zakresu
            st.session_state.prawidlowe_zmiany = prawidlowe_zmiany
            st.session_state.y_vars = {
                (p, d, s, l): y[p, d, s, l].varValue
                for p in pracownicy
                for d in dni_zakresu
                for s, l in prawidlowe_zmiany
            }
            st.session_state.wymagani_h = wymagani_pracownicy_h

        except Exception as e:
            st.error(f"⚠️ Wystąpił błąd podczas obliczeń: {e}")

# --- 5. INTERAKTYWNY PODGLĄD, EDYTOR NA ŻYWO I ANALITYKA ---
if st.session_state.get("schedule_generated", False):
    st.divider()
    st.header("5. Podgląd Grafiku & Analityka Obsady DS")

    pracownicy = st.session_state.pracownicy
    dni_zakresu = st.session_state.dni_zakresu
    prawidlowe_zmiany = st.session_state.prawidlowe_zmiany
    y_vars = st.session_state.y_vars

    def format_time(h_float):
        h_int = int(h_float) % 24
        m_int = int(round((h_float - int(h_float)) * 60))
        return f"{h_int:02d}:{m_int:02d}"

    data_rows = []
    for d in dni_zakresu:
        row = {"Data": d.strftime("%d/%m/%Y"), "Dzień": MAPA_DNI.get(d.strftime("%A"), "")}
        for p in pracownicy:
            shift_str = "OFF"
            for s, l in prawidlowe_zmiany:
                if y_vars.get((p, d, s, l), 0) == 1:
                    shift_str = f"{format_time(s)} - {format_time(s + l)}"
                    break
            row[p] = shift_str
        data_rows.append(row)

    df_editor = pd.DataFrame(data_rows)

    st.subheader("📝 Edytuj grafik na żywo:")
    st.info("💡 Kliknij w dowolną komórkę, aby zmienić godziny pracy (np. '06:00 - 14:00' lub 'OFF'). Zmiany zaktualizują wykresy i plik Excel.")

    edited_df = st.data_editor(df_editor, num_rows="fixed", use_container_width=True)

    st.subheader("📊 Analityka Obsady i Godzin Pickerów")
    tab1, tab2 = st.tabs(["📈 Pokrycie Zamówień w Dobie", "⚖️ Suma Godzin Pickerów"])

    with tab1:
        selected_day_str = st.selectbox("Wybierz dzień do analizy:", [d.strftime("%d/%m/%Y") for d in dni_zakresu])
        sel_date = next(d for d in dni_zakresu if d.strftime("%d/%m/%Y") == selected_day_str)

        req_hours = st.session_state.wymagani_h.get(sel_date, {})
        actual_hours = {h: 0 for h in range(26)}

        for p in pracownicy:
            val = edited_df.loc[edited_df["Data"] == selected_day_str, p].values[0]
            if str(val).strip() != "OFF" and "-" in str(val):
                try:
                    parts = str(val).split("-")
                    h_s = float(parts[0].split(":")[0]) + float(parts[0].split(":")[1]) / 60.0
                    h_e = float(parts[1].split(":")[0]) + float(parts[1].split(":")[1]) / 60.0
                    if h_e < h_s:
                        h_e += 24.0
                    for h in range(int(h_s), int(h_e)):
                        actual_hours[h] += 1
                except:
                    pass

        chart_data = pd.DataFrame({
            "Godzina": [f"{h:02d}:00" for h in range(6, 24)],
            "Wymagana obsada (Looker)": [req_hours.get(h, 0) for h in range(6, 24)],
            "Grafikowana obsada": [actual_hours.get(h, 0) for h in range(6, 24)],
        }).set_index("Godzina")

        st.bar_chart(chart_data)

    with tab2:
        worker_totals = {p: 0.0 for p in pracownicy}
        for p in pracownicy:
            for _, r in edited_df.iterrows():
                val = r[p]
                if str(val).strip() != "OFF" and "-" in str(val):
                    try:
                        parts = str(val).split("-")
                        h_s = float(parts[0].split(":")[0]) + float(parts[0].split(":")[1]) / 60.0
                        h_e = float(parts[1].split(":")[0]) + float(parts[1].split(":")[1]) / 60.0
                        if h_e < h_s:
                            h_e += 24.0
                        worker_totals[p] += (h_e - h_s)
                    except:
                        pass

        df_totals = pd.DataFrame(list(worker_totals.items()), columns=["Picker", "Suma Godzin (RH)"]).set_index("Picker")
        st.bar_chart(df_totals)

    st.subheader("📥 Eksport do Pliku Excel")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grafik jush"
    ws.freeze_panes = "B3"

    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3"),
    )

    fill_header_main = PatternFill(start_color="005B2B", end_color="005B2B", fill_type="solid")
    font_header_main = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    fill_header_sub = PatternFill(start_color="8BC53F", end_color="8BC53F", fill_type="solid")
    font_header_sub = Font(name="Calibri", size=10, bold=True, color="005B2B")
    fill_summary = PatternFill(start_color="EBF7D4", end_color="EBF7D4", fill_type="solid")
    fill_total_sum = PatternFill(start_color="8BC53F", end_color="8BC53F", fill_type="solid")
    font_total_sum = Font(name="Calibri", size=11, bold=True, color="005B2B")
    fill_shift_morning = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_shift_afternoon = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")

    ws.merge_cells("A1:A2")
    ws["A1"] = "pl-waw-12"
    ws["A1"].font = font_header_main
    ws["A1"].fill = fill_header_main
    ws["A1"].alignment = align_center

    col_idx = 2
    for p in pracownicy:
        col_start_letter = openpyxl.utils.get_column_letter(col_idx)
        col_end_letter = openpyxl.utils.get_column_letter(col_idx + 2)
        ws.merge_cells(f"{col_start_letter}1:{col_end_letter}1")
        cell_p = ws[f"{col_start_letter}1"]
        cell_p.value = p
        cell_p.font = font_header_main
        cell_p.fill = fill_header_main
        cell_p.alignment = align_center

        for i, sh in enumerate(["Start", "Koniec", "Suma"]):
            cell_sh = ws.cell(row=2, column=col_idx + i)
            cell_sh.value = sh
            cell_sh.font = font_header_sub
            cell_sh.fill = fill_header_sub
            cell_sh.alignment = align_center
            cell_sh.border = thin_border
        col_idx += 3

    row_idx = 3
    for _, r in edited_df.iterrows():
        cell_date = ws.cell(row=row_idx, column=1)
        cell_date.value = r["Data"]
        cell_date.font = font_regular
        cell_date.alignment = align_center
        cell_date.border = thin_border

        col_idx = 2
        for p in pracownicy:
            val = str(r[p]).strip()
            if val != "OFF" and "-" in val:
                parts = val.split("-")
                c_start = ws.cell(row=row_idx, column=col_idx)
                c_end = ws.cell(row=row_idx, column=col_idx + 1)
                c_sum = ws.cell(row=row_idx, column=col_idx + 2)

                c_start.value = parts[0].strip()
                c_end.value = parts[1].strip()

                try:
                    h_s = float(parts[0].split(":")[0]) + float(parts[0].split(":")[1]) / 60.0
                    h_e = float(parts[1].split(":")[0]) + float(parts[1].split(":")[1]) / 60.0
                    if h_e < h_s:
                        h_e += 24.0
                    c_sum.value = round(h_e - h_s, 1)
                except:
                    c_sum.value = 0

                fill_c = fill_shift_morning if "06:00" in parts[0] else fill_shift_afternoon
                for cell in [c_start, c_end, c_sum]:
                    cell.font = font_regular
                    cell.alignment = align_center
                    cell.border = thin_border
                    cell.fill = fill_c
            else:
                for i in range(3):
                    ws.cell(row=row_idx, column=col_idx + i).border = thin_border
            col_idx += 3
        row_idx += 1

    ws.column_dimensions["A"].width = 16
    for c in range(2, col_idx):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 10

    buffer = io.BytesIO()
    wb.save(buffer)

    st.download_button(
        label="📥 Pobierz Gotowy Grafik Excel (.xlsx)",
        data=buffer.getvalue(),
        file_name="grafik_pickerzy_jush.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
